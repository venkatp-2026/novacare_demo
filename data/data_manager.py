"""Data manager for handling Excel-based data storage with default and working copies."""
import os
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class DataManager:
    """Manages data storage with default and working copies in Excel."""
    
    def __init__(self, excel_file: str = "data/novacare_data.xlsx"):
        self.excel_file = Path(excel_file)
        self.patients: Dict[str, Dict[str, Any]] = {}
        self.appointments: Dict[str, List[Dict[str, Any]]] = {}
        self.slots: Dict[str, Dict[str, Any]] = {}
        
    def initialize_excel_with_defaults(self) -> None:
        """Create Excel file with default data if it doesn't exist."""
        if self.excel_file.exists():
            return
        
        # Ensure directory exists
        self.excel_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook with two sheets
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create default and working sheets
        default_sheet = wb.create_sheet('default')
        working_sheet = wb.create_sheet('working')
        
        # Initialize default data
        default_data = self._get_default_data()
        
        # Write to both sheets
        for sheet in [default_sheet, working_sheet]:
            self._write_data_to_sheet(sheet, default_data)
        
        # Save workbook
        wb.save(self.excel_file)
        print(f"Created Excel file: {self.excel_file}")
    
    def _get_default_data(self) -> Dict[str, Any]:
        """Get the default seed data."""
        return {
            "patients": {
                "PAT-001": {
                    "patient_id": "PAT-001",
                    "name": "Jane Smith",
                    "dob": "1985-03-15",
                    "mrn": "MRN-2001",
                    "status": "active"
                },
                "PAT-002": {
                    "patient_id": "PAT-002",
                    "name": "Robert Johnson",
                    "dob": "1978-11-22",
                    "mrn": "MRN-2002",
                    "status": "active"
                },
                "PAT-003": {
                    "patient_id": "PAT-003",
                    "name": "Maria Garcia",
                    "dob": "1992-07-08",
                    "mrn": "MRN-2003",
                    "status": "active"
                },
            },
            "appointments": {
                "PAT-001": [
                    {
                        "appointment_id": "APT-101",
                        "patient_id": "PAT-001",
                        "date": "2026-06-20",
                        "time": "10:00 AM",
                        "provider": "Dr. Williams",
                        "type": "Follow-up",
                        "location": "Main Campus - Room 204",
                        "status": "confirmed"
                    },
                    {
                        "appointment_id": "APT-102",
                        "patient_id": "PAT-001",
                        "date": "2026-07-15",
                        "time": "2:30 PM",
                        "provider": "Dr. Chen",
                        "type": "Annual Check-up",
                        "location": "North Clinic - Suite 301",
                        "status": "confirmed"
                    },
                    {
                        "appointment_id": "APT-103",
                        "patient_id": "PAT-001",
                        "date": "2026-08-10",
                        "time": "9:00 AM",
                        "provider": "Dr. Williams",
                        "type": "Lab Results Review",
                        "location": "Main Campus - Room 204",
                        "status": "confirmed"
                    }
                ],
                "PAT-002": [
                    {
                        "appointment_id": "APT-201",
                        "patient_id": "PAT-002",
                        "date": "2026-06-22",
                        "time": "11:30 AM",
                        "provider": "Dr. Patel",
                        "type": "Consultation",
                        "location": "West Building - Office 102",
                        "status": "confirmed"
                    },
                    {
                        "appointment_id": "APT-202",
                        "patient_id": "PAT-002",
                        "date": "2026-07-05",
                        "time": "3:00 PM",
                        "provider": "Dr. Lee",
                        "type": "Physical Therapy",
                        "location": "Therapy Center - Room A",
                        "status": "confirmed"
                    }
                ],
                "PAT-003": [
                    {
                        "appointment_id": "APT-301",
                        "patient_id": "PAT-003",
                        "date": "2026-06-25",
                        "time": "1:00 PM",
                        "provider": "Dr. Rodriguez",
                        "type": "Diabetes Management",
                        "location": "Endocrinology Wing - Room 5",
                        "status": "confirmed"
                    },
                    {
                        "appointment_id": "APT-302",
                        "patient_id": "PAT-003",
                        "date": "2026-07-20",
                        "time": "10:30 AM",
                        "provider": "Dr. Kim",
                        "type": "Nutrition Counseling",
                        "location": "Wellness Center - Suite 210",
                        "status": "confirmed"
                    }
                ]
            },
            "slots": {
                "SLOT-001": {"slot_id": "SLOT-001", "date": "2026-06-25", "time": "2:00 PM", "provider": "Dr. Williams", "location": "Main Campus - Room 204", "available": True},
                "SLOT-002": {"slot_id": "SLOT-002", "date": "2026-06-26", "time": "9:30 AM", "provider": "Dr. Chen", "location": "North Clinic - Suite 301", "available": True},
                "SLOT-003": {"slot_id": "SLOT-003", "date": "2026-06-27", "time": "11:00 AM", "provider": "Dr. Williams", "location": "Main Campus - Room 204", "available": True},
                "SLOT-004": {"slot_id": "SLOT-004", "date": "2026-06-28", "time": "3:30 PM", "provider": "Dr. Patel", "location": "West Building - Office 102", "available": True},
                "SLOT-005": {"slot_id": "SLOT-005", "date": "2026-07-01", "time": "8:00 AM", "provider": "Dr. Lee", "location": "Therapy Center - Room A", "available": True},
                "SLOT-006": {"slot_id": "SLOT-006", "date": "2026-07-02", "time": "1:30 PM", "provider": "Dr. Rodriguez", "location": "Endocrinology Wing - Room 5", "available": True},
                "SLOT-007": {"slot_id": "SLOT-007", "date": "2026-07-03", "time": "10:00 AM", "provider": "Dr. Kim", "location": "Wellness Center - Suite 210", "available": True},
                "SLOT-008": {"slot_id": "SLOT-008", "date": "2026-07-05", "time": "4:00 PM", "provider": "Dr. Williams", "location": "Main Campus - Room 204", "available": True},
                "SLOT-009": {"slot_id": "SLOT-009", "date": "2026-07-08", "time": "9:00 AM", "provider": "Dr. Chen", "location": "North Clinic - Suite 301", "available": True},
                "SLOT-010": {"slot_id": "SLOT-010", "date": "2026-07-10", "time": "2:30 PM", "provider": "Dr. Patel", "location": "West Building - Office 102", "available": True},
            }
        }
    
    def _write_data_to_sheet(self, sheet, data: Dict[str, Any]) -> None:
        """Write structured data to a sheet as JSON."""
        sheet['A1'] = 'Data Type'
        sheet['B1'] = 'JSON Data'
        
        row = 2
        for key, value in data.items():
            sheet[f'A{row}'] = key
            sheet[f'B{row}'] = json.dumps(value, ensure_ascii=False)
            row += 1
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 100
    
    def _read_data_from_sheet(self, sheet) -> Dict[str, Any]:
        """Read structured data from a sheet."""
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                data_type = row[0]
                json_data = row[1]
                try:
                    data[data_type] = json.loads(json_data)
                except json.JSONDecodeError:
                    print(f"Warning: Could not parse JSON for {data_type}")
        return data
    
    def load_from_working_copy(self) -> None:
        """Load data from working sheet into memory."""
        if not self.excel_file.exists():
            self.initialize_excel_with_defaults()
        
        wb = openpyxl.load_workbook(self.excel_file)
        
        if 'working' not in wb.sheetnames:
            raise ValueError("Working sheet not found in Excel file")
        
        working_sheet = wb['working']
        data = self._read_data_from_sheet(working_sheet)
        
        self.patients = data.get('patients', {})
        self.appointments = data.get('appointments', {})
        self.slots = data.get('slots', {})
        
        wb.close()
        print(f"Loaded data from working sheet: {len(self.patients)} patients, "
              f"{sum(len(appts) for appts in self.appointments.values())} appointments, "
              f"{len(self.slots)} slots")
    
    def save_to_working_copy(self) -> None:
        """Save current in-memory data to working sheet."""
        if not self.excel_file.exists():
            self.initialize_excel_with_defaults()
        
        wb = openpyxl.load_workbook(self.excel_file)
        
        if 'working' not in wb.sheetnames:
            raise ValueError("Working sheet not found in Excel file")
        
        working_sheet = wb['working']
        
        # Clear existing data
        for row in working_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        
        # Write current data
        data = {
            'patients': self.patients,
            'appointments': self.appointments,
            'slots': self.slots
        }
        self._write_data_to_sheet(working_sheet, data)
        
        wb.save(self.excel_file)
        wb.close()
        print("Saved data to working sheet")
    
    def refresh_from_default(self) -> None:
        """Reset working data from default sheet."""
        if not self.excel_file.exists():
            self.initialize_excel_with_defaults()
        
        wb = openpyxl.load_workbook(self.excel_file)
        
        if 'default' not in wb.sheetnames or 'working' not in wb.sheetnames:
            raise ValueError("Required sheets not found in Excel file")
        
        default_sheet = wb['default']
        working_sheet = wb['working']
        
        # Read from default
        data = self._read_data_from_sheet(default_sheet)
        
        # Clear working sheet
        for row in working_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        
        # Write to working
        self._write_data_to_sheet(working_sheet, data)
        
        wb.save(self.excel_file)
        wb.close()
        
        # Reload into memory
        self.load_from_working_copy()
        print("Refreshed working data from default")


# Global instance
_data_manager: Optional[DataManager] = None


def get_data_manager() -> DataManager:
    """Get the global data manager instance."""
    global _data_manager
    if _data_manager is None:
        _data_manager = DataManager()
    return _data_manager


def load_data_on_startup() -> None:
    """Initialize and load data on application startup."""
    dm = get_data_manager()
    dm.initialize_excel_with_defaults()
    dm.load_from_working_copy()


def save_working_data() -> None:
    """Save current data to working sheet."""
    dm = get_data_manager()
    dm.save_to_working_copy()
